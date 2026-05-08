import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

FILE_PATH = "/home/aman/Desktop/attendance_python_automation_project/dataset/attendance_2023_2024.csv"

print("Reading data .....")

df = pd.read_csv(FILE_PATH)


def calculate_time(time_data):
    try:
        if pd.isna(time_data):
            return 0

        login, logout = time_data.split("-")

        # calculate time
        login_time = datetime.strptime(login, "%H:%M")
        logout_time = datetime.strptime(logout, "%H:%M")

        # numbers of hours worked
        total_time = logout_time - login_time

        # Convert times into hours
        hours = total_time.seconds / 3600

        return round(hours, 2)
    except:
        return 0


print("Calculating Working hours....")
working_report = pd.DataFrame()

working_report["Date"] = df["Date"]
for employee in df.columns[1:]:
    working_report[employee] = df[employee].apply(calculate_time)

print("Checking low attendance")
low_attendance = pd.DataFrame()

# add date column
low_attendance["Date"] = working_report["Date"]

# Condition
MINIMUM_HOURS = 8

for employee in working_report.columns[1:]:
    low_attendance[employee] = working_report[employee].apply(
        lambda x: "LOW ATTENDANCE" if x < MINIMUM_HOURS else "OK"
    )


# Create summary report
print("Creating Summary Report...")
summary = []
for employee in working_report.columns[1:]:
    total_days = len(working_report)
    low_day = (low_attendance[employee] == "LOW ATTENDANCE").sum()
    good_day = (low_attendance[employee] == "OK").sum()
    attendance_percentage = (good_day / total_days) * 100

    summary.append(
        {
            "Employee": employee,
            "Total Day": total_days,
            "Present Day": good_day,
            "Absent Day": low_day,
            "Attendance Percentage": attendance_percentage,
        }
    )

summary_df = pd.DataFrame(summary)

# Save to excel
output_file = "Professional_attendance_report.xlsx"
print("Saving Excel")

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    working_report.to_excel(writer, sheet_name="Working hours", index=False)
    low_attendance.to_excel(writer, sheet_name="Low Attendance", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

print("Applying formatting")

wb = load_workbook(output_file)

ws = wb["Low Attendance"]

# Red color
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Green color
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

bold_font = Font(bold=True)

# Make header bold
for cell in ws[1]:
    cell.font = bold_font

# high light low attendace
for row in ws.iter_rows(min_row=2):
    for cell in row:
        if cell.value == "Low Attendance":
            cell.fill = red_fill
        elif cell.value == "OK":
            cell.fill = green_fill

# Auto adjust
for sheet in wb.sheetnames:
    work_sheet = wb[sheet]

    for column in work_sheet.columns:
        max_lenght = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(str(cell.value))
            except:
                pass
        adjusted_width = max_lenght + 5

        work_sheet.column_dimensions[column_letter].width = adjusted_width


# Save work book
wb.save(output_file)

print(f"Done report => {output_file}")
